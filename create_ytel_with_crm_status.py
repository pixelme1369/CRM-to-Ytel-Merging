import csv
import re
import sys
from pathlib import Path
from urllib.parse import unquote

from openpyxl import Workbook, load_workbook


def normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[-10:] if len(digits) >= 10 else ""


def resolve_csv_path(csv_path_or_folder):
    path = Path(csv_path_or_folder)
    if path.is_file():
        return path
    if not path.is_dir():
        raise FileNotFoundError(f"CRM path does not exist: {path}")

    csv_files = [
        file
        for file in path.glob("*.csv")
        if file.is_file() and not file.name.startswith(".")
    ]
    if not csv_files:
        raise FileNotFoundError(f"No .csv files found in CRM folder: {path}")

    return max(csv_files, key=lambda file: file.stat().st_mtime)


def resolve_ytel_path(ytel_path_or_folder):
    path = Path(ytel_path_or_folder)
    if path.is_file():
        return path
    if not path.is_dir():
        raise FileNotFoundError(f"Ytel path does not exist: {path}")

    ytel_files = [
        file
        for pattern in ("*.xlsx", "*.csv")
        for file in path.glob(pattern)
        if file.is_file() and not file.name.startswith(".") and not file.name.startswith("~$")
    ]
    if not ytel_files:
        raise FileNotFoundError(f"No .xlsx or .csv files found in Ytel folder: {path}")

    return max(ytel_files, key=lambda file: file.stat().st_mtime)


def load_ytel_workbook(ytel_path):
    if ytel_path.suffix.lower() == ".xlsx":
        return load_workbook(ytel_path)
    if ytel_path.suffix.lower() == ".csv":
        wb = Workbook()
        ws = wb.active
        ws.title = ytel_path.stem[:31] or "Ytel"
        with open(ytel_path, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                ws.append(row)
        return wb
    raise ValueError(f"Unsupported Ytel file type: {ytel_path}")


def clean_status(value):
    status = str(value or "").strip()
    status = re.sub(r'[\/\\:*?"<>|]', "-", status)
    status = re.sub(r"\s+", "_", status)
    status = re.sub(r"-+", "-", status)
    return status


def phone_from_recording(value):
    text = unquote(str(value or ""))
    matches = re.findall(r"_(\d{10,})(?=\D|$)", text)
    if matches:
        return normalize_phone(matches[-1])
    runs = re.findall(r"\d{10,}", text)
    return normalize_phone(runs[-1]) if runs else ""


def append_status_to_recording(value, status):
    text = str(value or "")
    if not text or not status:
        return text

    suffix = f"-{status}"
    match = re.search(r"(\.[A-Za-z0-9]{2,5})(\?.*)?$", text)
    if match:
        base = text[: match.start()]
        ext_and_query = text[match.start() :]
        if base.endswith(suffix):
            return text
        return f"{base}{suffix}{ext_and_query}"

    if text.endswith(suffix):
        return text
    return f"{text}{suffix}"


EXTRA_CRM_COLUMNS = [
    "Enrolled Debt",
    "Last Credit Pulled Date",
    "Cordoba Enrolled Date",
    "Credit Score",
    "Assigned To",
    "State",
]

# (crm_column, output_column, required)
CRM_IMPORT_COLUMNS = [
    ("Enrolled Debt", "Enrolled Debt", True),
    ("Last Credit Pulled Date", "Last Credit Pulled Date", True),
    ("Cordoba Enrolled Date", "Cordoba Enrolled Date", True),
    ("Credit Score", "Credit Score", True),
    ("ID", "AMOD", True),
    ("Assigned To", "Assigned To", False),
    ("State", "State", False),
]

REMOVE_OUTPUT_COLUMNS = {
    "call_list_id",
    "vendor_lead_code",
    "list_id",
    "gmt_offset_now",
    "phone_code",
    "title",
    "gender",
    "date_of_birth",
    "alt_phone",
    "email",
    "security_phrase",
    "comments",
    "list_name",
    "owner",
    "postal_code",
    "rank",
    "country_code",
    "province",
    "city",
    "address3",
    "address2",
    "address1",
    "middle_initial",
}


def read_crm_rows_by_phone(csv_path, home_phone_col, status_col):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CRM CSV has no header row.")

        field_map = {name.strip().lower(): name for name in reader.fieldnames}
        home_key = field_map.get(home_phone_col.strip().lower())
        status_key = field_map.get(status_col.strip().lower())
        extra_keys = {}
        for crm_column, output_column, required in CRM_IMPORT_COLUMNS:
            key = field_map.get(crm_column.strip().lower())
            if not key:
                if required:
                    raise ValueError(f'Could not find CRM column "{crm_column}". Found: {reader.fieldnames}')
                print(f'Warning: optional CRM column "{crm_column}" not found — "{output_column}" will be blank.', file=sys.stderr)
                continue
            extra_keys[output_column] = key

        if not home_key:
            raise ValueError(f'Could not find CRM column "{home_phone_col}". Found: {reader.fieldnames}')
        if not status_key:
            raise ValueError(f'Could not find CRM column "{status_col}". Found: {reader.fieldnames}')

        crm_by_phone = {}
        rows = 0
        for row in reader:
            rows += 1
            phone = normalize_phone(row.get(home_key))
            status = clean_status(row.get(status_key))
            if phone and phone not in crm_by_phone:
                crm_by_phone[phone] = {
                    "CRM Status": status,
                    **{output_column: row.get(key, "") for output_column, key in extra_keys.items()},
                }
        return crm_by_phone, rows


def find_header_column(ws, header_name):
    target = header_name.strip().lower()
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if str(value or "").strip().lower() == target:
            return col_idx
    return None


def get_headers(ws):
    return {
        str(ws.cell(row=1, column=col_idx).value or "").strip().lower(): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }


def find_ytel_phone_and_recording(ws, row_idx, output_headers):
    headers = get_headers(ws)
    phone_number_dialed_col = headers.get("phone_number_dialed")
    if phone_number_dialed_col:
        phone = normalize_phone(ws.cell(row=row_idx, column=phone_number_dialed_col).value)
        if phone:
            recording_value = ""
            for col_idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
                if header in output_headers:
                    continue
                value = ws.cell(row=row_idx, column=col_idx).value
                text = str(value or "")
                if "_" in text or ".mp3" in text.lower() or ".txt" in text.lower():
                    recording_value = text
                    break
            return phone, recording_value

    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
        if header in output_headers:
            continue
        value = ws.cell(row=row_idx, column=col_idx).value
        text = str(value or "")
        phone = phone_from_recording(value)
        if phone and ("_" in text or ".mp3" in text.lower() or ".txt" in text.lower()):
            return phone, text

    for header in ["phone_number", "alt_phone"]:
        col_idx = headers.get(header)
        if not col_idx:
            continue
        phone = normalize_phone(ws.cell(row=row_idx, column=col_idx).value)
        if phone:
            return phone, ""

    return "", ""


def remove_duplicate_output_columns(ws, headers):
    seen = set()
    for col_idx in range(ws.max_column, 0, -1):
        value = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
        if value in headers:
            if value in seen:
                ws.delete_cols(col_idx)
            else:
                seen.add(value)


def remove_unwanted_output_columns(ws):
    for col_idx in range(ws.max_column, 0, -1):
        header = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
        if header in REMOVE_OUTPUT_COLUMNS:
            ws.delete_cols(col_idx)


def main():
    if len(sys.argv) < 4:
        print(
            "Usage: python create_ytel_with_crm_status.py <crm.csv-or-folder> <ytel.xlsx/csv-or-folder> <output.xlsx> "
            "[home-phone-column] [status-column]",
            file=sys.stderr,
        )
        return 1

    crm_csv = resolve_csv_path(sys.argv[1])
    ytel_path = resolve_ytel_path(sys.argv[2])
    output_xlsx = Path(sys.argv[3])
    home_phone_col = sys.argv[4] if len(sys.argv) > 4 else "Home Phone"
    status_col = sys.argv[5] if len(sys.argv) > 5 else "Status"

    crm_by_phone, crm_rows = read_crm_rows_by_phone(crm_csv, home_phone_col, status_col)

    wb = load_ytel_workbook(ytel_path)
    ws = wb.active

    crm_output_columns = [output_column for _, output_column, _ in CRM_IMPORT_COLUMNS]
    output_column_names = ["CRM Status", *crm_output_columns, "Recording With CRM Status"]
    output_headers = {name.lower() for name in output_column_names}
    remove_duplicate_output_columns(ws, output_headers)

    output_cols = {}
    for header in output_column_names:
        col_idx = find_header_column(ws, header)
        if not col_idx:
            col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx).value = header
        output_cols[header] = col_idx

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in output_cols.values():
            ws.cell(row=row_idx, column=col_idx).value = None

    matched = 0
    changed = 0

    for row_idx in range(2, ws.max_row + 1):
        recording_phone, recording_value = find_ytel_phone_and_recording(ws, row_idx, output_headers)

        if not recording_phone:
            continue

        crm_row = crm_by_phone.get(recording_phone)
        if not crm_row:
            continue

        matched += 1
        for header in ["CRM Status", *crm_output_columns]:
            ws.cell(row=row_idx, column=output_cols[header]).value = crm_row.get(header, "")

        if recording_value:
            updated_recording = append_status_to_recording(recording_value, crm_row.get("CRM Status", ""))
        else:
            updated_recording = f"{recording_phone}-{crm_row.get('CRM Status', '')}".rstrip("-")
        ws.cell(row=row_idx, column=output_cols["Recording With CRM Status"]).value = updated_recording

        if recording_value and updated_recording != recording_value:
            changed += 1

    remove_unwanted_output_columns(ws)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(
        {
            "crm_rows": crm_rows,
            "crm_source": str(crm_csv),
            "ytel_source": str(ytel_path),
            "crm_phones_with_status": len(crm_by_phone),
            "ytel_rows": max(ws.max_row - 1, 0),
            "matched_ytel_rows": matched,
            "recording_names_with_status_added": changed,
            "output": str(output_xlsx),
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
